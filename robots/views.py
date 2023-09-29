from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from django.http import HttpResponse
from openpyxl import Workbook
from .models import Robot
from datetime import datetime, timedelta
from collections import defaultdict

@csrf_exempt
def create_robot(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            model = data.get('model')
            version = data.get('version')
            created_str = data.get('created')
            
            
            if not model or len(model) != 2:
                return JsonResponse({"error": "Invalid model"}, status=400)
            if not version or len(version) != 2:
                return JsonResponse({"error": "Invalid version"}, status=400)
            if not created_str:
                return JsonResponse({"error": "Invalid created"}, status=400)
            
            try:
                created = datetime.strptime(created_str, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                return JsonResponse({"error": "Invalid date format"}, status=400)

            serial = f"{model}-{version}"
            
            Robot.objects.create(serial=serial, model=model, version=version, created=created)
            
            return JsonResponse({"success": True}, status=201)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


def export_excel(request):
    wb = Workbook()
    wb.remove(wb.active)

    one_week_ago = datetime.now() - timedelta(weeks=1)
    robots = Robot.objects.filter(created__gte=one_week_ago)

    grouped_data = defaultdict(list)
    
    for robot in robots:
        grouped_data[robot.model].append(robot)        
   
    for model, robots_list in grouped_data.items():
        ws = wb.create_sheet(title=model)       
        ws.append(["Модель", "Версия", "Количество за неделю"])
        
        version_count = defaultdict(int)
        for robot in robots_list:
            version_count[robot.version] += 1
        for version, count in version_count.items():
            ws.append([model, version, count])
            
    if not wb.worksheets or all(ws.sheet_state == 'hidden' for ws in wb.worksheets):
        wb.create_sheet(title="Empty")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Robots_report.xlsx'
    wb.save(response)

    return response